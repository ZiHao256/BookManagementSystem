"""系统运行前的准备程序"""
"""Books.xlsx:第一行为书名，第二行为作者，第三行为出版社，第四行为价格，第五行为借阅情况（是/否），第六行为借阅人，第七行为借阅时间"""
""""Student.xlsx:学生注册后，sheet名为学生学号，第一列为学生信息:姓名，学号。第二列：第一行为借阅书籍：第一行为书名，第二行为是否是否已还，第三行为借阅时间，第四行为归还时间"""
""""Librarians.xlsx:管理员注册后，sheet名为管理员学工号，第一列为管理员信息：姓名，学工号，密码；第二列为操作记录"""
import openpyxl
Books = openpyxl.Workbook("Books.xlsx")
Books.save("Books.xlsx")
print("Books库创建完成...")
Students = openpyxl.Workbook("Students.xlsx")
Students.save("Students.xlsx")
print("Students库创建完成...")
Librarians = openpyxl.Workbook("Librarians.xlsx")
Librarians.save("Librarians.xlsx")
print("Librarians库创建完成...\n图书管理系统可以正常工作！")