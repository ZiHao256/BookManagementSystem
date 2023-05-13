"""图书管理系统正式运行的程序"""
import openpyxl
import sys
import datetime

def InitialInterface():
    """
    初始界面：
    1.登陆
    2.注册
    3.退出
    """
    print("***欢迎使用图书管理系统***")
    print('1.登录账号')
    print('2.注册账号')
    print('3.退出')
    x = int(input())
    if x == 1:
        SignIn()
    if x == 2:
        Register()
    if x == 3:
        print("再见！")
        sys.exit(0)  # 执行该语句会直接退出程序，这也是经常使用的方法，也不需要考虑平台等因素的影响，一般是退出Python程序的首选方法
def Register():
    """
    注册页面：
    1.学生注册
    2.管理员注册
    3.返回
    """
    print('1.学生注册')
    print('2.管理员注册')
    print('3.返回')
    x = int(input())
    if x == 1:
        name = input("输入姓名：")
        number = input("输入学号：")
        a_student = Students(name, number)
        a_student.SRegisiter()
    if x == 2:
        name = input("输入姓名：")
        number = input("输入学工号：")
        key = input("输入密码：")
        a_librarian = Librarians(name, number, key)
        a_librarian.LRegister()
    if x == 3:
        InitialInterface()
def SignIn():
    """
    登陆界面：
    1.学生登陆
    2.管理员登陆
    3.返回主页面
    """
    print('1.学生登陆')
    print('2.管理员登陆')
    print('3.返回')
    x = int(input())
    if x == 1:
        name = input("输入姓名：")
        number = input("输入学号：")
        a_student = Students(name, number)
        a_student.SSignIn()
    if x == 2:
        name = input("输入姓名：")
        number = input("输入学工号：")
        key = input("输入密码：")
        a_librarian = Librarians(name,number,key)
        a_librarian.LSignIn()
    if x == 3:
        InitialInterface()
def LOperation(a_liabrarian):
    # 登陆成功后的管理员操作
    print("1.添加图书\n2.查找图书\n3.归还图书\n4.退出")
    x = int(input())

    if x == 1:
        a_liabrarian.LAddBooks()
    if x == 2:
        a_liabrarian.LSearchBooks()
    if x == 3:
        a_liabrarian.LReturnBooks()
    if x == 4:
        a_liabrarian.LQuit()
def SOperation(a_student):
    # 登陆成功后的学生操作
    print("1.查询图书\n2.归还图书\n3.退出")
    x = int(input())
    if x == 1:
        a_student.SSearchBooks()
    if x == 2:
        a_student.SReturnBooks()
    if x == 3:
        print("再见！")
        a_student.SQuit()
class Students(object):
    # 学生管理(姓名和学号）
    # 1.查询图书2.借出图书3.归还图书4.退出程序

    def __init__(self, name, number):
        self.name = name
        self.number = number
    def SRegisiter(self):
        """"
            学生注册：
            1.若注册过，则返回注册页面
            2.若未注册，则将学生各项信息记录入相应表内
        """
        Workbook = openpyxl.load_workbook("Students.xlsx")
        Registered = 1
        while(Registered):
            try:
                Worksheet = Workbook[self.number]
                print("已注册！ \n请重新选择注册")
                Register()
                Registered = 1
            except:
                Worksheet = Workbook.create_sheet(self.number)
                Registered = 0


        #在学生表内记录学生的各项信息：姓名，学号；借阅书籍：书名，是否已还，借阅时间，归还时间
        Worksheet.cell(1,1).value = self.name
        Worksheet.cell(2,1).value = self.number
        Worksheet.cell(1,2).value = "借阅书籍"
        Worksheet.cell(2,2).value = "是否已还"
        Worksheet.cell(3,2).value = "借阅时间"
        Worksheet.cell(4,2).value = "归还时间"
        Workbook.save("Students.xlsx")
        print("注册成功！\n请重新登录")
        SignIn()
    def SSignIn(self):
        """"
        学生登陆：
            1.将输入的学号与已注册的学生学号对比
            2.若已存在，则对比姓名：若姓名不匹配，则输出学号与姓名不匹配
                                若姓名匹配，则登陆成功
            3。若学号不存在，则返回主页面
        """
        Workbook = openpyxl.load_workbook("Students.xlsx")
        fail = 1                                         #判断是否匹配
        while(fail):
            for i in Workbook.sheetnames:
                if i == self.number:
                    Worksheet = Workbook[self.number]
                    if Worksheet.cell(1,1).value == self.name:  #匹配
                        print("登陆成功！")
                        fail = 0
            if fail == 1:                                    #不匹配
                print("登陆失败！姓名与学号不匹配！\n请重新输入")
                SignIn()
        # 登陆成功 #
        SOperation(self)
    def SSearchBooks(self):
        """"在相应种类里查找并输出书的各项信息："""
        WorkBook = openpyxl.load_workbook("Books.xlsx")
        ans = 1                                                # 判断是否有该种类
        while (ans): # 查找书的种类
            print("馆内书的种类：%s" %(WorkBook.sheetnames))
            BookType = input("输入书的种类：")
            try:
                WorkSheet = WorkBook[BookType]
                ans = 0
            except:
                print("无该类书，请重新输入")
                ans = 1
        ans2 = 1  # 判断是否有该书
        while (ans2):
            BookName = input("输入书名：")
            col = WorkSheet.max_column
            for i in list(range(2, col + 1)):

                if WorkSheet.cell(1, i).value == BookName:      # 有该书
                    ans2 = 0
                    print("书名：%s" % (WorkSheet.cell(1, i).value))
                    print("作者：%s" % (WorkSheet.cell(2, i).value))
                    print("出版社：%s" % (WorkSheet.cell(3, i).value))
                    print("价格：%s" % (WorkSheet.cell(4, i).value))
                    print("借阅情况：%s" % (WorkSheet.cell(5, i).value))

            if ans2 == 1:
                print("无该书，重新输入")
        x = int(input("1.借阅该图书\n2.返回\n3.退出"))
        if x == 1:
            self.SBorrowBooks(BookType, i)
        if x == 2:
            SOperation(self)
        if x == 3:
            self.SQuit()
    def SBorrowBooks(self, Type, col):
        """
        学生借书：
        1.对书的搜索将书的种类及该书的列数输入该函数
        2.对书进行借阅：1、将被借阅的书的信息（借阅情况，借阅人，借阅时间）补充完整
                     2、将借阅同学的借阅历史（借阅书籍、是否已还、借阅时间、归还时间）补充完整
        :return:
        """
        Workbook = openpyxl.load_workbook("Books.xlsx")
        BookType = Type
        WorkSheet = Workbook[BookType]
        BookName = WorkSheet.cell(1,col).value
                # 借阅操作
        x = int(input("确定借阅%s ?\n1.借阅\n2.返回" %(WorkSheet.cell(1,col).value)))            # intput()为str类型
        if x == 1:
                    # 将被借阅的书信息补充完整
            cur_time = datetime.datetime.now()
            WorkSheet.cell(5, col).value = "是"         # 借阅情况
            WorkSheet.cell(6, col).value = self.name   # 借阅人
            WorkSheet.cell(7, col).value = cur_time    # 借阅时间
            Workbook.save("Books.xlsx")
                    # 将该同学的借阅历史补充完整
            StuWb = openpyxl.load_workbook("Students.xlsx")
            StuWs = StuWb[self.number]
            col = StuWs.max_column+1                  # 存储位置
            StuWs.cell(1, col).value = BookName
            StuWs.cell(2, col).value = "否"
            StuWs.cell(3, col).value = cur_time
            StuWs.cell(4, col).value = "无"
            StuWb.save("Students.xlsx")             # 总是忘记储存xlsx！！！！！！！
            print("借阅成功！")
        if x == 2:
            SOperation(self)
        SOperation(self)
    def SReturnBooks(self):
        """"
                学生还书：
                1.首先对Books.xlsx操作：找到书籍对应种类，找到书列，将书籍的信息更新
                2.然后对Students,xlsx操作：找到借阅同学，找到借阅历史对应的一列,修改借阅信息
                3.Books.xlsx和Students保存
        """
        BookWb = openpyxl.load_workbook("Books.xlsx")
        print("馆内书的种类：%s" %(BookWb.sheetnames))
        BookType = input("输入归还书籍的种类：")
        BookName = input("输入归还书籍名称：")
        BorrowerNumber = self.number
        # 对Books.xlsx更新
        BookWb = openpyxl.load_workbook("Books.xlsx")
        BookWs = BookWb[BookType]
        for i in list(range(2, BookWs.max_column + 1)):
            if BookWs.cell(1, i).value == BookName:
                BookWs.cell(5, i).value = "否"
                BookWs.cell(6, i).value = "无"
                BookWs.cell(7, i).value = "无"
        BookWb.save("Books.xlsx")
        # 对Students.xlsx更新
        cur_time = datetime.datetime.now()
        StuWb = openpyxl.load_workbook("Students.xlsx")
        StuWs = StuWb[BorrowerNumber]
        for i in list(range(3, StuWs.max_column + 1)):
            if StuWs.cell(1, i).value == BookName:
                StuWs.cell(2, i).value = "是"
                StuWs.cell(4, i).value = cur_time
        StuWb.save("Students.xlsx")
        print("归还成功！")
        SOperation(self)
    def SQuit(self):
        print("再见！")
        sys.exit(0)
class Librarians(object):
    # 管理员管理
    # 1.新增图书2.查询图书3.修改图书4.删除图书4.归还图书5.退出程序
    # 管理员由于管理能力大，都需要记录操作（添加、删除操作需要记录
    def __init__(self,name,number,key):
        self.number = number
        self.name = name
        self.key = key
    def LSignIn(self):
        """
                管理员登陆：
                    1.将输入的学工号与已注册的学生学号对比
                    2.若已存在，则对比姓名与密码：若姓名不匹配，则输出学号与姓名不匹配：若密码不匹配，则输出密码错误
                                            若姓名匹配且密码正确，则登陆成功
                    3。若学工号不存在，则返回主页面
                """
        Workbook = openpyxl.load_workbook("Librarians.xlsx")
        fail = 1  # 判断是否匹配
        while (fail):
            for i in Workbook.sheetnames:
                if i == self.number:
                    Worksheet = Workbook[self.number]
                    if (Worksheet.cell(1, 1).value == self.name) & (Worksheet.cell(3, 1).value == self.key):  # 姓名和密码均匹配
                        print("登陆成功！")
                        fail = 0
                        break
            if fail == 1:  # 不匹配
                print("登陆失败！姓名与学工号不匹配或者密码错误！\n请重新输入")
                SignIn()
        # 登陆成功：可以进行管理员的操作 #
        LOperation(self)
    def LRegister(self):
        """"
        管理员注册：
            1.若注册过，则返回注册页面
            2.若未注册，则将管理员信息记录
        """
        Workbook = openpyxl.load_workbook("Librarians.xlsx")
        Registered = 1
        while (Registered):
            try:
                Worksheet = Workbook[self.number]
                print("该学工号已被注册！ \n请重新选择注册")
                Register()
                Registered = 1
            except:
                Worksheet = Workbook.create_sheet(self.number)
                Registered = 0

        # 在管理员表内记录管理员的各项信息：姓名，学工号,密码；对图书的各项操作，及时间
        Worksheet.cell(1, 1).value = self.name
        Worksheet.cell(2, 1).value = self.number
        Worksheet.cell(3, 1).value = self.key
        Worksheet.cell(1, 2).value = "添加图书"
        Worksheet.cell(2, 2).value = "添加时间"
        Worksheet.cell(3, 2).value = "删除图书"
        Worksheet.cell(4, 2).value = "删除时间"
        Workbook.save("Librarians.xlsx")
        print("注册成功！\n请重新登陆")
        SignIn()
    def LAddBooks(self):
        """"根据书的种类将书的各项信息存入Books.xlsx"""
        Workbook = openpyxl.load_workbook("Books.xlsx")              #打开Books.xlxs
        BookItems = []
        print("已有书的种类：%s" %(Workbook.sheetnames))
        BookType = input("输入书的种类:")
        BookName = input("输入书名:")
        BookAuthor = input("输入作者:")
        BookPublishment = input("输入出版社:")
        BookPrice = input("输入书价:")
        BookItems = [BookName,BookAuthor,BookPublishment,BookPrice]  #用列表存储，方便输入数据

        try:                                                         #将书放入相应的种类，若无则创建
            Worksheet = Workbook[BookType]
        except:
            Worksheet = Workbook.create_sheet(BookType)
            for i in list(range(1,8)):
                Worksheet.cell(i, 1).value = ["书名", "作者", "出版社", "书价", "借阅情况", "借阅人", "借阅时间"][i-1]

        col = Worksheet.max_column + 1                               #已存书目的个数+1
        Worksheet.cell(1, col).value = BookItems[0]                   #书名
        Worksheet.cell(2, col).value = BookItems[1]                   #作者
        Worksheet.cell(3, col).value = BookItems[2]                   #出版社
        Worksheet.cell(4, col).value = BookItems[3]                   #书价
        Worksheet.cell(5, col).value = "否"                           #借阅情况
        Worksheet.cell(6, col).value = "无"                            #借阅人
        Worksheet.cell(7, col).value = "无"                            #借阅时间
        Workbook.save("Books.xlsx")
        print("添加成功！")
        # 将该操作添加入该管理员的操作历史
        cur_time = datetime.datetime.now()
        LibWb = openpyxl.load_workbook("Librarians.xlsx")
        LibWs = LibWb[self.number]
        col = LibWs.max_column
        LibWs.cell(1, col+1).value = BookName
        LibWs.cell(2, col+1).value = cur_time
        LibWb.save("Librarians.xlsx")
        x = int(input("1.返回\n2.退出"))
        if x == 1:
            LOperation(self)
        if x == 2:
            self.LQuit()
    def LSearchBooks(self):
        """"在相应种类里查找并输出书的各项信息："""
        WorkBook = openpyxl.load_workbook("Books.xlsx")
        ans = 1                                         #判断是否有该种类
        while(ans):                                     #查找书的种类
            print("馆内书的种类：%s" %(WorkBook.sheetnames))
            BookType = input("输入书的种类：")
            try:
                WorkSheet = WorkBook[BookType]
                ans =0
            except:
                print("无该类书，请重新输入")
                ans = 1
        ans2 = 1                                         #判断是否有该书
        while(ans2):
            BookName = input("输入书名：")
            col = WorkSheet.max_column
            for i in list(range(2,col+1)):

                if WorkSheet.cell(1,i).value == BookName:#有该书
                    ans2 = 0
                    print("书名：%s" %(WorkSheet.cell(1,i).value))
                    print("作者：%s" % (WorkSheet.cell(2, i).value))
                    print("出版社：%s" % (WorkSheet.cell(3, i).value))
                    print("价格：%s" % (WorkSheet.cell(4, i).value))
                    print("借阅情况：%s" %(WorkSheet.cell(5,i).value))
                    if WorkSheet.cell(5,i).value == "是":
                        print("借阅人：%s" %(WorkSheet.cell(6,i).value))
                        print("借阅时间：%s" %(WorkSheet.cell(7,i).value))

            if ans2 == 1:
                print("无该书！")
                x = int(input("1.重新输入\n2.返回"))
                if x == 1:
                    continue
                if x == 2:
                    self.LSearchBooks()
        print("1.删除该图书\n2.修改图书信息\n3.返回")
        x = int(input())
        if x == 1:
            y = int(input("确定删除%s?\n1.确定\n2.返回" %(WorkSheet.cell(1,i).value)))
            if y == 1:
                # 更新管理员操作历史
                cur_time = datetime.datetime.now()
                LibWb = openpyxl.load_workbook("Librarians.xlsx")
                LibWs = LibWb[self.number]
                col = LibWs.max_column
                LibWs.cell(3, col+1).value = BookName
                LibWs.cell(4, col+1).value = cur_time
                LibWb.save("Librarians.xlsx")  # 又又又忘记保存
                # 更新书库信息
                WorkSheet.delete_cols(i)
                WorkBook.save("Books.xlsx")  # 又忘记保存！！！！

                print("删除成功！")
                LOperation(self)
            if y == 2:
                LOperation(self)
        if x == 2:
            y = int(input("确定修改%s的信息?\n1.确定\n2.返回" % (WorkSheet.cell(1, i).value)))
            if y == 1:
                print("书名：%s" % (WorkSheet.cell(1, i).value))
                print("作者：%s" % (WorkSheet.cell(2, i).value))
                print("出版社：%s" % (WorkSheet.cell(3, i).value))
                print("价格：%s" % (WorkSheet.cell(4, i).value))
                print("借阅情况：%s" % (WorkSheet.cell(5, i).value))
                print("借阅人：%s" % (WorkSheet.cell(6, i).value))
                print("借阅时间：%s" % (WorkSheet.cell(7, i).value))
                z = int(input("1.修改书名\n2.修改作者\n3.修改出版社\n4.修改价格\n5.修改借阅情况\n6.修改借阅人\n7.修改借阅时间"))
                change = input("修改为：")
                if z == 1:
                    WorkSheet.cell(1,i).value = change
                if z == 2:
                    WorkSheet.cell(2,i).value = change
                if z == 3:
                    WorkSheet.cell(3,i).value = change
                if z == 4:
                    WorkSheet.cell(4,i).value = change
                if z == 5:
                    WorkSheet.cell(5,i).value = change
                if z == 6:
                    WorkSheet.cell(6,i).value = change
                if z == 7:
                    WorkSheet.cell(7,i).value = change
                WorkBook.save("Books.xlsx")
                print("修改成功！")
                LOperation(self)

            if y == 2:
                LOperation(self)
        if x == 3:
            LOperation(self)
    def LBorrowBooks(self):
        pass
    def LReturnBooks(self):
        """"
        管理员还书：
        1.首先对Books.xlsx操作：找到书籍对应种类，找到书列，将书籍的信息更新
        2.然后对Students,xlsx操作：找到借阅同学，找到借阅历史对应的一列,修改借阅信息
        3.Books.xlsx和Students保存
        """
        BookWb = openpyxl.load_workbook("Books.xlsx")
        print("馆内书的种类：%s" %(BookWb.sheetnames))
        BookType = input("输入归还书籍的种类：")
        BookName = input("输入归还书籍名称：")
        BorrowerNumber = input("输入归还学生学号：")
        # 对Books.xlsx更新

        BookWs = BookWb[BookType]
        for i in list(range(2, BookWs.max_column+1)):
            if BookWs.cell(1, i).value == BookName:
                BookWs.cell(5, i).value = "否"
                BookWs.cell(6, i).value = "无"
                BookWs.cell(7, i).value = "无"
        BookWb.save("Books.xlsx")
        # 对Students.xlsx更新
        cur_time = datetime.datetime.now()
        StuWb = openpyxl.load_workbook("Students.xlsx")
        StuWs = StuWb[BorrowerNumber]
        for i in list(range(3,StuWs.max_column+1)):
            if StuWs.cell(1, i).value == BookName:
                StuWs.cell(2, i).value = "是"
                StuWs.cell(4, i).value = cur_time
        StuWb.save("Students.xlsx")
        print("归还成功！")
        x = int(input("1.返回\n2.退出"))
        if x == 1:
            LOperation(self)
        if x == 2:
            self.LQuit()
    def LQuit(self):
        print("再见！")
        sys.exit(0)
# 主程序
InitialInterface()
